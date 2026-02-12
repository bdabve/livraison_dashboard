#!/usr/bin/env python3
# -*- coding: utf-8 -*-
#
# author        : el3arbi bdabve@gmail.com
# created       :
# desc          :
# ----------------------------------------------------------------------------
# import time
import os
import pandas as pd
import undetected_chromedriver as uc

from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException     # , StaleElementReferenceException

# ---------- CONFIG SETUP ----------
BASE_URL = "http://51.255.79.241:8080/trizstock"
LOGIN_URL = f"{BASE_URL}/faces/login.xhtml"
DEFAULT_TIMEOUT = 20
PRODUIT_SORTIE_URL = f"{BASE_URL}/faces/view/livraison/listDetailProduitSortieN.xhtml"


# ---------- DRIVER SETUP ----------
def create_driver(headless: bool = False, chrome_version=143, download_dir="./triz_downloads"):
    """
    :return: driver
    """
    os.makedirs(download_dir, exist_ok=True)    # Create download dir if it doesn't exist
    download_dir = os.path.abspath(download_dir)
    options = uc.ChromeOptions()
    prefs = {
        "download.default_directory": download_dir,
        "download.prompt_for_download": False,
        "download_directory_upgrade": True,
        "safebrowsing.enabled": True
    }
    options.add_experimental_option('prefs', prefs)
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument(f"--unsafely-treat-insecure-origin-as-secure={BASE_URL}")

    return uc.Chrome(version_main=chrome_version, options=options, headless=headless)


# ---------- LOGIN ----------
def login(driver, username: str, password: str, timeout: int = DEFAULT_TIMEOUT) -> bool:
    print("[+] Login...")

    wait = WebDriverWait(driver, timeout)
    driver.get(LOGIN_URL)

    # Wait until inputs are visible
    user_input = wait.until(EC.visibility_of_element_located((By.ID, "j_username")))
    pass_input = wait.until(EC.visibility_of_element_located((By.ID, "j_password")))

    user_input.clear()
    user_input.send_keys(username)

    pass_input.clear()
    pass_input.send_keys(password)
    pass_input.send_keys(Keys.ENTER)

    # ---- wait for either success OR error (shorter wait) ----
    short_wait = WebDriverWait(driver, 5)

    try:
        error_box = short_wait.until(
            EC.visibility_of_element_located((By.ID, "errorMessages"))
        )
        msg = error_box.find_element(By.CLASS_NAME, "ui-messages-error-summary").text
        return {"success": False, "message": f"[✗] Login failed: {msg}"}

    except TimeoutException:
        return {"success": True, "message": "[✓] Login successful."}


# ---------- DOWNLOAD EXCEL FILE ----------
def download_etat_prevendeur(driver, dated, datef, camion):
    """
    This function will download the Excel File
    """
    wait = WebDriverWait(driver, DEFAULT_TIMEOUT)
    url = (
        f"{PRODUIT_SORTIE_URL}?"
        f"{'camion=' + camion + '&' if camion else ''}"
        f"datef={datef}&dated={dated}&statut=livrer"
    )
    driver.get(url)
    wait.until(EC.presence_of_element_located((By.TAG_NAME, "body")))
    excel_btn = driver.find_element(By.ID, 'j_idt545:j_idt604:j_idt610')
    excel_btn.click()
    input("Press Enter after the download is complete...")
    print("[✓] Excel download initiated.")


# ---------- DOWNLOAD EXCEL FOR PREVENDEUR ----------
def download_all_etats(driver, dated, datef):
    camions = {
        "VENTE": "",        # All
        "WALID": "8442-0000005",
        "MOHAMED": "8442-0000006",
        "FETHI": "8442-0000007",
        "MM": "8442-0000010"
    }
    for prev, camion in camions.items():
        print('-' * 40)
        print(f"[+] Download Excel for {prev}.")
        download_etat_prevendeur(driver, dated, datef, camion=camion)


def merge_excels_with_sheetnames(input_folder, output_file):
    excel_files = sorted([
        f for f in os.listdir(input_folder)
        if f.endswith((".xlsx", ".xls"))
    ])
    # Merge all files into a single Excel with sheet names prevendeur
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        for file in excel_files:
            file_path = os.path.join(input_folder, file)
            sheet_name = os.path.splitext(file)[0]

            print(f"Adding {file} -> sheet: {sheet_name}")

            df = pd.read_excel(file_path)
            df.to_excel(writer, sheet_name=sheet_name[:31], index=False)

    print("Done ✅", output_file)


from datetime import date, datetime


def faresse_etat(driver):
    base_url = "http://51.255.79.241:8080/trizstock/faces/view/livraison/list.xhtml"

    camions = {
        "WALID": "8442-0000005",
        "MOHAMED": "8442-0000006",
        "FETHI": "8442-0000007",
        "MM": "8442-0000010"
    }

    today = date.today()

    # First day of current month
    start_date = today.replace(day=1)

    results = []

    current = start_date
    while current <= today:

        # Format for URL (dd-mm-yyyy)
        formatted_date = current.strftime("%d-%m-%Y")

        daily_totals = {}

        for prev, camion in camions.items():
            url = (
                f"{base_url}?"
                f"{'camion=' + camion + '&' if camion else ''}"
                f"datef={formatted_date}&dated={formatted_date}&statut=livrer"
            )

            driver.get(url)

            total_livrer = driver.find_element(
                By.XPATH,
                '//*[@id="listeLiv:j_idt663:dataTable_foot"]/tr/td[7]'
            )

            clean_value = (
                total_livrer.text
                .replace('\u202f', '')
                .replace('\u00a0', '')
                .replace(' ', '')
                .replace(',', '.')
            )

            daily_totals[prev] = float(clean_value)

        results.append({
            formatted_date: daily_totals
        })

        current = current.replace(day=current.day + 1)

    return results


def export_faress_etat_excel(data, filename="faresse_etat.xlsx"):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Etat Livraison"

    # Extract camion names dynamically
    first_day = next(iter(data[0].values()))
    camions = list(first_day.keys())

    # Header row
    headers = ["Date"] + camions
    ws.append(headers)

    # Data rows
    for day_data in data:
        date_str = next(iter(day_data.keys()))
        totals = day_data[date_str]

        row = [date_str] + [totals.get(camion, 0) for camion in camions]
        ws.append(row)

    wb.save(filename)
    print(f"✅ Excel file created: {filename}")


if __name__ == '__main__':
    import dotenv
    dotenv.load_dotenv(dotenv.find_dotenv())
    username = os.getenv("triz_username")
    passwd = os.getenv('triz_password')

    driver = create_driver()
    result = login(driver, username, passwd)
    # if result["success"]:
    #     # # ------
    #     print(result["message"])
    #     dated = "01-01-2026"
    #     datef = "31-01-2026"
    #     download_all_etats(driver, dated, datef)
    # else:
    #     print(result["message"])

    # input_folder = "./triz_downloads"
    # output_file = "VENTE_JANVIER_2026.xlsx"
    # # sheet_names = ["WALID", "MOHAMED", "FETHI", "MM", "VENTE"]
    # merge_excels_with_sheetnames(input_folder, output_file)

    # FARESSE ETAT
    if result["success"]:
        etat_faress = faresse_etat(driver)
        export_faress_etat_excel(etat_faress)
        print(etat_faress)
